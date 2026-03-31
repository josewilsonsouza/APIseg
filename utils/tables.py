"""
utils/tables.py
Helpers de formatação e construção de tabelas de resumo. Os export
"""

import pandas as pd
from ui.colors import CNSEG_TEAL, CNSEG_RED
from utils.formatting import (
    _br,
    prev_month as _prev_month,
    short_lbl as _short_lbl,
    fmt_bi,
    fmt_pct,
)


def var_pct(atual: float, anterior: float) -> float | None:
    if anterior and anterior != 0:
        return (atual - anterior) / abs(anterior) * 100
    return None


def make_table(
    src: pd.DataFrame,
    col: str,
    periodo_ref: int,
    ref_label: str,
    periodo_ant: int = None,
) -> pd.DataFrame:
    """
    Constrói tabela de resumo por segmento com R$ bi, M/M (%) e A/A (%).
    Funciona tanto para dados SUSEP quanto Cognos, desde que a coluna
    de agrupamento seja 'segmento'.
    """
    cur = src[src["damesano"] == periodo_ref][["segmento", col]].copy()
    cur = cur.rename(columns={col: "atual", "segmento": "Segmento"}).set_index("Segmento")
    cur[f"R$ bi ({ref_label})"] = cur["atual"] / 1e9

    if periodo_ant and periodo_ant in src["damesano"].values:
        ant = src[src["damesano"] == periodo_ant].set_index("segmento")[[col]]
        ant.index.name = "Segmento"
        cur = cur.join(ant.rename(columns={col: "_ant"}))
        cur["M/M (%)"] = (
            (cur["atual"] - cur["_ant"]) / cur["_ant"].abs().replace(0, float("nan")) * 100
        )

    yoy_period = periodo_ref - 100
    if yoy_period in src["damesano"].values:
        yoy = src[src["damesano"] == yoy_period].set_index("segmento")[[col]]
        yoy.index.name = "Segmento"
        cur = cur.join(yoy.rename(columns={col: "_yoy"}))
        cur["A/A (%)"] = (
            (cur["atual"] - cur["_yoy"]) / cur["_yoy"].abs().replace(0, float("nan")) * 100
        )

    out = [f"R$ bi ({ref_label})"] + [c for c in ["M/M (%)", "A/A (%)"] if c in cur.columns]
    tbl = cur[out].sort_values(f"R$ bi ({ref_label})", ascending=False)
    tbl.loc["TOTAL"] = {f"R$ bi ({ref_label})": tbl[f"R$ bi ({ref_label})"].sum()}
    return tbl.reset_index()


def calc_12mm(
    df_full: pd.DataFrame,
    col: str,
    group_col: str = "segmento",
) -> pd.DataFrame:
    """
    Calcula soma rolling 12 meses para cada grupo.
    Retorna DataFrame com as mesmas colunas de damesano/periodo_label/group_col
    e a coluna `col` substituída pelo valor acumulado dos últimos 12 períodos.
    """
    df = df_full.sort_values(["damesano"]).copy()
    result_rows = []
    for grp, grp_df in df.groupby(group_col, sort=False):
        grp_df = grp_df.sort_values("damesano").reset_index(drop=True)
        grp_df[col] = grp_df[col].rolling(12, min_periods=1).sum()
        result_rows.append(grp_df)
    return pd.concat(result_rows, ignore_index=True)


def make_wide_table(
    df_full: pd.DataFrame,
    col: str,
    n_months: int = 24,
    divisor: float = 1e9,
    group_col: str = "segmento",
) -> pd.DataFrame:
    """
    Tabela histórica: rows=grupo, cols=últimos n_months períodos (R$ bi).
    """
    periodos = sorted(df_full["damesano"].unique())[-n_months:]
    df = df_full[df_full["damesano"].isin(periodos)].copy()
    df["_val"] = df[col] / divisor
    pivot = df.pivot_table(
        index=group_col, columns="periodo_label", values="_val", aggfunc="sum"
    )
    # Mantém colunas ordenadas cronologicamente
    ordered_cols = (
        df[["damesano", "periodo_label"]]
        .drop_duplicates()
        .sort_values("damesano")["periodo_label"]
        .tolist()
    )
    ordered_cols = [c for c in ordered_cols if c in pivot.columns]
    pivot = pivot[ordered_cols]
    pivot.index.name = "Segmento" if group_col == "segmento" else group_col.capitalize()
    return pivot.reset_index()


# ---------------------------------------------------------------------------
# Hierarchical table helpers
# ---------------------------------------------------------------------------

_HIER_COLS = ["segmento", "grupo1", "grupo2", "ramo"]


def _hier_agg(df: pd.DataFrame, col: str, hier_cols: list) -> dict:
    """Pre-compute groupby sums at each hierarchy level. Returns {level: Series}."""
    result = {}
    for i in range(len(hier_cols)):
        keys = hier_cols[: i + 1]
        if all(k in df.columns for k in keys):
            result[i] = df.groupby(keys, observed=True)[col].sum(min_count=1)
    return result


def _hier_val(aggs: dict, level: int, key) -> float:
    """Look up an aggregated value; returns NaN on miss."""
    s = aggs.get(level)
    if s is None or s.empty:
        return float("nan")
    k = key[0] if isinstance(key, tuple) and len(key) == 1 else key
    try:
        return float(s.loc[k])
    except (KeyError, TypeError):
        return float("nan")

def make_hierarchical_table(
    df_raw: pd.DataFrame,
    col: str,
    periodo_ref: int,
    ref_label: str,
    periodo_ant: int = None,
    denom_col: str = None,
) -> pd.DataFrame:
    """
    Constrói tabela hierárquica segmento -> grupo1 -> grupo2 -> ramo.

    Se denom_col for fornecido, exibe col/denom_col*100 como sinistralidade (%).
    Retorna DataFrame com colunas: Nome, valor, M/M, A/A, _level.
    O campo _level deve ser ocultado via style_hierarchical() ao exibir.
    """
    needed = ["damesano"] + _HIER_COLS + [col]
    if denom_col:
        needed.append(denom_col)
    df = df_raw[[c for c in needed if c in df_raw.columns]].copy()
    df[col] = pd.to_numeric(df[col], errors="coerce")

    hier_cols = [c for c in _HIER_COLS if c in df.columns]
    is_ratio = bool(denom_col and denom_col in df.columns)

    if is_ratio:
        df[denom_col] = pd.to_numeric(df[denom_col], errors="coerce")

    yoy_period = periodo_ref - 100
    cur_df = df[df["damesano"] == periodo_ref]
    ant_df = (
        df[df["damesano"] == periodo_ant]
        if periodo_ant and periodo_ant in df["damesano"].values
        else None
    )
    yoy_df = df[df["damesano"] == yoy_period] if yoy_period in df["damesano"].values else None

    aggs_cur = _hier_agg(cur_df, col, hier_cols)
    aggs_ant = _hier_agg(ant_df, col, hier_cols) if ant_df is not None else {}
    aggs_yoy = _hier_agg(yoy_df, col, hier_cols) if yoy_df is not None else {}

    if is_ratio:
        den_cur = _hier_agg(cur_df, denom_col, hier_cols)
        den_ant = _hier_agg(ant_df, denom_col, hier_cols) if ant_df is not None else {}
        den_yoy = _hier_agg(yoy_df, denom_col, hier_cols) if yoy_df is not None else {}

    # Column names
    if is_ratio:
        val_col = f"Sinistral. ({ref_label}) %"
        mm_col = "M/M (p.p.)"
        aa_col = "A/A (p.p.)"
    else:
        val_col = f"R$ bi ({ref_label})"
        mm_col = "M/M (%)"
        aa_col = "A/A (%)"

    def _ratio(num, den):
        return num / den * 100 if not (pd.isna(num) or pd.isna(den) or den == 0) else float("nan")

    def make_row(label: str, level: int, key) -> dict:
        num_c = _hier_val(aggs_cur, level, key)
        num_a = _hier_val(aggs_ant, level, key)
        num_y = _hier_val(aggs_yoy, level, key)

        if is_ratio:
            v = _ratio(num_c, _hier_val(den_cur, level, key))
            v_a = _ratio(num_a, _hier_val(den_ant, level, key))
            v_y = _ratio(num_y, _hier_val(den_yoy, level, key))
            mm = (v - v_a) if not (pd.isna(v) or pd.isna(v_a)) else float("nan")
            aa = (v - v_y) if not (pd.isna(v) or pd.isna(v_y)) else float("nan")
        else:
            v = num_c / 1e9 if not pd.isna(num_c) else float("nan")
            mm = (
                (num_c - num_a) / abs(num_a) * 100
                if not (pd.isna(num_c) or pd.isna(num_a) or num_a == 0)
                else float("nan")
            )
            aa = (
                (num_c - num_y) / abs(num_y) * 100
                if not (pd.isna(num_c) or pd.isna(num_y) or num_y == 0)
                else float("nan")
            )
        return {"Nome": label, "_level": level, val_col: v, mm_col: mm, aa_col: aa}

    _SP = "\u00a0\u00a0\u00a0"  # indent unit (3 non-breaking spaces per level)

    rows: list[dict] = []
    n = len(hier_cols)

    for seg in sorted(cur_df["segmento"].dropna().unique()):
        rows.append(make_row(seg, 0, seg))
        if n < 2:
            continue

        g1s = sorted(cur_df.loc[cur_df["segmento"] == seg, "grupo1"].dropna().unique())
        for g1 in g1s:
            show_g1 = str(g1) != str(seg)
            if show_g1:
                rows.append(make_row(_SP + str(g1), 1, (seg, g1)))
            if n < 3:
                continue

            mask1 = (cur_df["segmento"] == seg) & (cur_df["grupo1"] == g1)
            g2s = sorted(cur_df.loc[mask1, "grupo2"].dropna().unique())
            for g2 in g2s:
                show_g2 = str(g2) != str(g1)
                if show_g2:
                    indent2 = _SP * 2 if show_g1 else _SP
                    rows.append(make_row(indent2 + str(g2), 2, (seg, g1, g2)))
                if n < 4:
                    continue

                mask2 = mask1 & (cur_df["grupo2"] == g2)
                for ramo in sorted(cur_df.loc[mask2, "ramo"].dropna().unique()):
                    if str(ramo) == str(g2):
                        continue
                    n_ind = show_g1 + show_g2 + 1
                    rows.append(make_row(_SP * n_ind + str(ramo), 3, (seg, g1, g2, ramo)))

    result = pd.DataFrame(rows)
    keep = ["Nome", val_col, "_level", mm_col, aa_col]
    keep = [c for c in keep if c in result.columns]
    return result[keep].reset_index(drop=True)


def style_hierarchical(df: pd.DataFrame, higher_is_worse: bool = False):
    """
    Aplica formatação e estilos a tabela hierárquica.
    Oculta a coluna _level e aplica cor/negrito por nível de hierarquia.
    higher_is_worse=True inverte as cores M/M e A/A (ex.: sinistralidade).
    """
    pct_cols = [c for c in df.columns if ("M/M" in c or "A/A" in c) and "%" in c]
    pp_cols = [c for c in df.columns if "p.p." in c]
    bi_cols = [c for c in df.columns if "R$ bi" in c]
    sr_cols = [c for c in df.columns if "Sinistral." in c]

    fmt: dict = {}
    fmt.update({c: (lambda v: _br(v, ".2f")) for c in bi_cols})
    fmt.update({c: (lambda v: _br(v, ".1f") + "%") for c in sr_cols})
    fmt.update({c: (lambda v: _br(v, "+.1f") + "%") for c in pct_cols})
    fmt.update({c: (lambda v: _br(v, "+.1f")) for c in pp_cols})

    def color_pct(val):
        if pd.isna(val):
            return ""
        good = val >= 0 if not higher_is_worse else val <= 0
        return (
            f"color: {CNSEG_TEAL}; font-weight:500"
            if good
            else f"color: {CNSEG_RED}; font-weight:500"
        )

    def row_style(row):
        level = int(row["_level"]) if "_level" in row.index and pd.notna(row["_level"]) else 0
        n = len(row)
        if level == 0:
            return ["font-weight:bold; background-color:rgba(247,135,31,0.08)"] * n
        if level == 1:
            return ["font-weight:600"] * n
        if level == 3:
            return ["color:#A09FA4; font-size:0.9em"] * n
        return [""] * n

    styler = df.style.format(fmt, na_rep="—").apply(row_style, axis=1)
    for c in pct_cols + pp_cols:
        styler = styler.map(color_pct, subset=[c])
    if "_level" in df.columns:
        styler = styler.hide(["_level"], axis="columns")
    return styler


def style_table(tbl: pd.DataFrame):
    """Aplica formatação e colorização de variações percentuais."""
    pct_cols = [c for c in tbl.columns if "%" in c]
    bi_cols = [c for c in tbl.columns if "R$ bi" in c]
    fmt = {c: (lambda v: _br(v, ".2f")) for c in bi_cols}
    fmt.update({c: (lambda v: _br(v, "+.1f") + "%") for c in pct_cols})

    def color(val):
        if pd.isna(val):
            return ""
        return (
            f"color: {CNSEG_TEAL}; font-weight:500"
            if val >= 0
            else f"color: {CNSEG_RED}; font-weight:500"
        )

    styler = tbl.style.format(fmt, na_rep="—")
    for c in pct_cols:
        styler = styler.map(color, subset=[c])
    return styler



def make_relatorio_mensal(
    df_raw: pd.DataFrame,
    col: str,
    periodo_ref: int,
    periodo_ant: int,
    periodo_labels: dict,
) -> pd.DataFrame:
    """
    Constrói tabela estilo relatório Excel com:
    - 5 colunas mensais em R$ mi
    - 2 variações M/M e 2 variações A/A
    - Acumulado YTD (ano anterior e ano ref) em R$ mi
    - Variação % acumulada YoY
    - Participação % no crescimento acumulado
    - Linha IPCA ao final
    """
    p_ref     = periodo_ref
    p_ant     = periodo_ant
    p_ant2    = _prev_month(p_ant)
    p_yoy     = p_ref - 100
    p_yoy_ant = p_ant - 100

    five_periods = [p_yoy_ant, p_yoy, p_ant2, p_ant, p_ref]

    df = df_raw.copy()
    df[col] = pd.to_numeric(df[col], errors="coerce")
    hier_cols = [c for c in _HIER_COLS if c in df.columns]

    all_periods = sorted(df["damesano"].dropna().unique())
    year_ref   = p_ref // 100
    year_prev  = year_ref - 1
    month_ref  = p_ref % 100

    acum_ref_periods  = [p for p in all_periods if p // 100 == year_ref  and p % 100 <= month_ref]
    acum_prev_periods = [p for p in all_periods if p // 100 == year_prev and p % 100 <= month_ref]

    def _agg_p(period: int) -> dict:
        sub = df[df["damesano"] == period]
        return _hier_agg(sub, col, hier_cols) if not sub.empty else {}

    def _agg_ps(periods: list) -> dict:
        sub = df[df["damesano"].isin(periods)]
        return _hier_agg(sub, col, hier_cols) if not sub.empty else {}

    aggs         = {p: _agg_p(p) for p in five_periods}
    agg_acum_ref  = _agg_ps(acum_ref_periods)
    agg_acum_prev = _agg_ps(acum_prev_periods)

    # Totais para participação no crescimento
    df_acum_ref  = df[df["damesano"].isin(acum_ref_periods)][col]
    df_acum_prev = df[df["damesano"].isin(acum_prev_periods)][col]
    total_a_ref  = df_acum_ref.sum(min_count=1)
    total_a_prev = df_acum_prev.sum(min_count=1)
    total_growth = (
        (total_a_ref - total_a_prev)
        if not (pd.isna(total_a_ref) or pd.isna(total_a_prev))
        else float("nan")
    )

    # IPCA por período (mesmo valor para todas as linhas do período)
    ipca_vals: dict = {}
    if "ipca" in df.columns:
        for p in five_periods:
            sub = df[df["damesano"] == p]["ipca"].dropna()
            if not sub.empty:
                ipca_vals[p] = sub.iloc[0]
        for tag, periods in [("acum_ref", acum_ref_periods), ("acum_prev", acum_prev_periods)]:
            sub = df[df["damesano"].isin(periods)]["ipca"].dropna()
            if not sub.empty:
                ipca_vals[tag] = sub.sum()  # IPCA acumulado = produto simplificado via soma

    def _lbl(p: int) -> str:
        return _short_lbl(periodo_labels.get(p, str(p)))

    m1, m2, m3, m4, m5 = [_lbl(p) for p in five_periods]
    mm_ref_col   = f"% {m5}/{m4}"
    mm_ant_col   = f"% {m4}/{m3}"
    aa_ref_col   = f"% {m5}/{m2}"
    aa_ant_col   = f"% {m4}/{m1}"
    acum_prev_col = str(year_prev)
    acum_ref_col  = str(year_ref)
    var_acum_col  = f"% {year_ref}/{year_prev}"
    part_col      = "Part. % crescimento"

    def _var(a, b):
        return (a - b) / abs(b) * 100 if not (pd.isna(a) or pd.isna(b) or b == 0) else float("nan")

    def _part(a_ref, a_prev):
        if pd.isna(total_growth) or total_growth == 0:
            return float("nan")
        g = (a_ref - a_prev) if not (pd.isna(a_ref) or pd.isna(a_prev)) else float("nan")
        return g / total_growth * 100 if not pd.isna(g) else float("nan")

    def make_row(label: str, level: int, key) -> dict:
        row: dict = {"Nome": label, "_level": level}
        vals = {p: _hier_val(aggs[p], level, key) for p in five_periods}
        for p in five_periods:
            row[_lbl(p)] = vals[p] / 1e6 if not pd.isna(vals[p]) else float("nan")

        row[mm_ref_col] = _var(vals[p_ref], vals[p_ant])
        row[mm_ant_col] = _var(vals[p_ant], vals[p_ant2])
        row[aa_ref_col] = _var(vals[p_ref], vals[p_yoy])
        row[aa_ant_col] = _var(vals[p_ant], vals[p_yoy_ant])

        a_ref  = _hier_val(agg_acum_ref,  level, key)
        a_prev = _hier_val(agg_acum_prev, level, key)
        row[acum_prev_col] = a_prev / 1e6 if not pd.isna(a_prev) else float("nan")
        row[acum_ref_col]  = a_ref  / 1e6 if not pd.isna(a_ref)  else float("nan")
        row[var_acum_col]  = _var(a_ref, a_prev)
        row[part_col]      = _part(a_ref, a_prev)
        return row

    # Itera mantendo a ordem natural dos dados (não alfabética)
    _SP = "\u00a0\u00a0\u00a0"
    rows: list[dict] = []
    n = len(hier_cols)

    cur_df = df[df["damesano"] == p_ref]
    seen_segs: list = []
    for seg in cur_df["segmento"].dropna():
        if seg not in seen_segs:
            seen_segs.append(seg)

    for seg in seen_segs:
        rows.append(make_row(seg, 0, seg))
        if n < 2:
            continue
        seg_df = cur_df[cur_df["segmento"] == seg]
        seen_g1: list = []
        for g1 in seg_df["grupo1"].dropna():
            if g1 not in seen_g1:
                seen_g1.append(g1)
        for g1 in seen_g1:
            show_g1 = str(g1) != str(seg)
            if show_g1:
                rows.append(make_row(_SP + str(g1), 1, (seg, g1)))
            if n < 3:
                continue
            g1_df = seg_df[seg_df["grupo1"] == g1]
            seen_g2: list = []
            for g2 in g1_df["grupo2"].dropna():
                if g2 not in seen_g2:
                    seen_g2.append(g2)
            for g2 in seen_g2:
                show_g2 = str(g2) != str(g1)
                if show_g2:
                    indent2 = _SP * 2 if show_g1 else _SP
                    rows.append(make_row(indent2 + str(g2), 2, (seg, g1, g2)))
                if n < 4:
                    continue
                g2_df = g1_df[g1_df["grupo2"] == g2]
                seen_ramo: list = []
                for ramo in g2_df["ramo"].dropna():
                    if ramo not in seen_ramo:
                        seen_ramo.append(ramo)
                for ramo in seen_ramo:
                    if str(ramo) == str(g2):
                        continue
                    n_ind = show_g1 + show_g2 + 1
                    rows.append(make_row(_SP * n_ind + str(ramo), 3, (seg, g1, g2, ramo)))

    # Linha de total
    total_row: dict = {"Nome": f"Setor Segurador (sem Saúde)", "_level": 0}
    for p in five_periods:
        v = df[df["damesano"] == p][col].sum(min_count=1)
        total_row[_lbl(p)] = v / 1e6 if not pd.isna(v) else float("nan")
    v_tot_ref  = df[df["damesano"] == p_ref][col].sum(min_count=1)
    v_tot_ant  = df[df["damesano"] == p_ant][col].sum(min_count=1)
    v_tot_ant2 = df[df["damesano"] == p_ant2][col].sum(min_count=1)
    v_tot_yoy  = df[df["damesano"] == p_yoy][col].sum(min_count=1)
    v_tot_yoy_ant = df[df["damesano"] == p_yoy_ant][col].sum(min_count=1)
    total_row[mm_ref_col]   = _var(v_tot_ref, v_tot_ant)
    total_row[mm_ant_col]   = _var(v_tot_ant, v_tot_ant2)
    total_row[aa_ref_col]   = _var(v_tot_ref, v_tot_yoy)
    total_row[aa_ant_col]   = _var(v_tot_ant, v_tot_yoy_ant)
    total_row[acum_prev_col] = total_a_prev / 1e6 if not pd.isna(total_a_prev) else float("nan")
    total_row[acum_ref_col]  = total_a_ref  / 1e6 if not pd.isna(total_a_ref)  else float("nan")
    total_row[var_acum_col]  = _var(total_a_ref, total_a_prev)
    total_row[part_col]      = 100.0
    rows.append(total_row)

    # Linha IPCA
    ipca_row: dict = {c: float("nan") for c in rows[0].keys()}
    ipca_row["Nome"] = "IPCA"
    ipca_row["_level"] = -1
    for p in five_periods:
        if p in ipca_vals:
            ipca_row[_lbl(p)] = ipca_vals[p]
    if "acum_ref" in ipca_vals:
        ipca_row[var_acum_col] = ipca_vals["acum_ref"]
    rows.append(ipca_row)

    result = pd.DataFrame(rows).reset_index(drop=True)

    # Extrai levels antes de montar MultiIndex
    levels: list[int] = result["_level"].fillna(0).astype(int).tolist()
    names: list[str]  = result["Nome"].tolist()
    result = result.drop(columns=["_level"])

    # Grupos para MultiIndex de colunas
    _G_MEN  = "Mensal (R$ mi)"
    _G_VAR  = "Var. Nominal"
    _G_ACUM = "Acumulado (R$ mi)"
    _G_CRES = "Crescimento Acum."

    mi_tuples = (
        [("Segmento", "")]
        + [(_G_MEN,  _lbl(p)) for p in five_periods]
        + [(_G_VAR,  mm_ref_col), (_G_VAR,  mm_ant_col),
           (_G_VAR,  aa_ref_col), (_G_VAR,  aa_ant_col)]
        + [(_G_ACUM, acum_prev_col), (_G_ACUM, acum_ref_col)]
        + [(_G_CRES, var_acum_col),  (_G_CRES, part_col)]
    )
    result.columns = pd.MultiIndex.from_tuples(mi_tuples)

    return result, levels, names


def style_relatorio_mensal(
    df: pd.DataFrame,
    levels: list[int],
    names: list[str],
) -> object:
    """Aplica formatação à tabela de relatório mensal com MultiIndex de colunas."""
    from ui.colors import CNSEG_TEAL, CNSEG_RED, CNSEG_ORANGE

    _G_MEN  = "Mensal (R$ mi)"
    _G_VAR  = "Var. Nominal"
    _G_ACUM = "Acumulado (R$ mi)"
    _G_CRES = "Crescimento Acum."

    mon_cols = [c for c in df.columns if c[0] in (_G_MEN, _G_ACUM) and c[0] != "Segmento"]
    var_cols = [c for c in df.columns if c[0] == _G_VAR]
    cres_acum_col = [c for c in df.columns if c[0] == _G_CRES and c[1].startswith("%")]
    part_col_full = [c for c in df.columns if c[0] == _G_CRES and not c[1].startswith("%")]

    fmt: dict = {}
    fmt.update({c: (lambda v: _br(v, ",.1f")) for c in mon_cols})
    fmt.update({c: (lambda v: _br(v, "+.1f") + "%") for c in var_cols + cres_acum_col})
    fmt.update({c: (lambda v: _br(v, ".1f") + "%") for c in part_col_full})

    def color_pct(val):
        if pd.isna(val):
            return ""
        return (
            f"color: {CNSEG_TEAL}; font-weight:500"
            if val >= 0
            else f"color: {CNSEG_RED}; font-weight:500"
        )

    def row_style(row):
        idx = row.name
        level  = levels[idx] if idx < len(levels) else 0
        name   = names[idx]  if idx < len(names)  else ""
        n = len(row)
        if level == -1:  # IPCA
            return [f"color: {CNSEG_ORANGE}; font-style:italic; font-size:0.85em"] * n
        if name.startswith("Setor Segurador"):
            return ["font-weight:bold; background-color:rgba(247,135,31,0.12)"] * n
        if level == 0:
            return ["font-weight:bold; background-color:rgba(247,135,31,0.08)"] * n
        if level == 1:
            return ["font-weight:600"] * n
        if level == 3:
            return ["color:#A09FA4; font-size:0.9em"] * n
        return [""] * n

    color_cols = var_cols + cres_acum_col + part_col_full
    styler = df.style.format(fmt, na_rep="—").apply(row_style, axis=1)
    for c in color_cols:
        styler = styler.map(color_pct, subset=[c])
    return styler


def add_saude_to_evol(by_seg: pd.DataFrame, raw_dataseg: pd.DataFrame, col: str) -> pd.DataFrame:
    """
    Adiciona Saúde Suplementar (trimestral) ao by_seg para o gráfico de evolução.
    Meses não-trimestrais ficam NaN -> gap no gráfico (connectgaps=False).
    """
    saude = raw_dataseg[raw_dataseg["segmento"] == "Saúde Suplementar"]
    periodos = by_seg["damesano"].unique()
    saude = saude[saude["damesano"].isin(periodos)]
    if saude.empty:
        return by_seg
    saude_agg = (
        saude.groupby(["damesano", "periodo_label", "segmento"])[col]
        .sum(min_count=1)
        .reset_index()
    )
    return pd.concat(
        [by_seg[["damesano", "periodo_label", "segmento", col]], saude_agg],
        ignore_index=True,
    )


def export_relatorio_excel(
    tbl: pd.DataFrame,
    levels: list,
    names: list,
    sheet_name: str = "Relatório",
) -> bytes:
    """
    Exporta make_relatorio_mensal para Excel com formatação completa:
    - Dois cabeçalhos (grupo + sub-coluna) com cores CNseg
    - Células mescladas no cabeçalho de grupo
    - Formatação numérica por tipo de coluna
    - Cores verde/vermelho nas variações
    - Larguras de coluna e freeze panes
    - Estilos por nível hierárquico
    """
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Paleta CNseg
    C_DARK   = "403D4A"
    C_DARK2  = "5A5761"
    C_ORANGE = "F7871F"
    C_ORANGE_BG  = "FEF3E8"
    C_ORANGE_BG2 = "FDE0B8"
    C_TEAL   = "20A787"
    C_RED    = "F44949"
    C_GRAY   = "A09FA4"
    WHITE    = "FFFFFF"

    buf = _io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    cols = tbl.columns.tolist()
    n_cols = len(cols)

    thin_side = Side(style="thin", color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Header row 1: grupos mesclados ──────────────────────────────────────
    fill1  = PatternFill(fill_type="solid", fgColor=C_DARK)
    font1  = Font(bold=True, color=WHITE, size=10)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Agrupa colunas consecutivas pelo primeiro nível do MultiIndex
    grp_spans: list[tuple[str, int, int]] = []
    prev_grp, start_i = None, 0
    for i, (g, _) in enumerate(cols):
        if g != prev_grp:
            if prev_grp is not None:
                grp_spans.append((prev_grp, start_i, i - 1))
            prev_grp, start_i = g, i
    grp_spans.append((prev_grp, start_i, len(cols) - 1))

    grp_border_side = Side(style="medium", color=WHITE)
    for grp, ci_start, ci_end in grp_spans:
        ec, sc = ci_end + 1, ci_start + 1          # openpyxl 1-based
        cell = ws.cell(row=1, column=sc, value=grp if grp != "Segmento" else "")
        cell.font = font1
        cell.fill = fill1
        cell.alignment = align_c
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        for c in range(sc, ec + 1):
            ws.cell(row=1, column=c).fill = fill1
        # Borda branca média nas extremidades do grupo
        ws.cell(row=1, column=sc).border = Border(left=grp_border_side)
        ws.cell(row=1, column=ec).border = Border(right=grp_border_side)

    # ── Header row 2: sub-colunas ────────────────────────────────────────────
    fill2 = PatternFill(fill_type="solid", fgColor=C_DARK2)
    font2 = Font(bold=True, color=WHITE, size=9)
    align_c2 = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (grp, sub) in enumerate(cols):
        cell = ws.cell(row=2, column=i + 1, value=sub if sub else grp)
        cell.font = font2
        cell.fill = fill2
        cell.alignment = align_c2
        cell.border = thin_border

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # ── Dados ────────────────────────────────────────────────────────────────
    _G_MEN  = "Mensal (R$ mi)"
    _G_ACUM = "Acumulado (R$ mi)"
    _G_VAR  = "Var. Nominal"
    _G_CRES = "Crescimento Acum."

    for row_idx, row_data in tbl.iterrows():
        xrow  = row_idx + 3
        level = levels[row_idx] if row_idx < len(levels) else 0
        name  = names[row_idx]  if row_idx < len(names)  else ""
        is_total = name.startswith("Setor Segurador")
        is_ipca  = (level == -1)

        # Escolhe fundo/fonte base por nível
        if is_ipca:
            base_fill = PatternFill(fill_type="solid", fgColor="FFF3E0")
            base_font = dict(italic=True, color=C_ORANGE, size=9)
        elif is_total:
            base_fill = PatternFill(fill_type="solid", fgColor=C_ORANGE_BG2)
            base_font = dict(bold=True, size=10)
        elif level == 0:
            base_fill = PatternFill(fill_type="solid", fgColor=C_ORANGE_BG)
            base_font = dict(bold=True, size=10)
        elif level == 1:
            base_fill = None
            base_font = dict(bold=True, size=9)
        elif level == 3:
            base_fill = None
            base_font = dict(color=C_GRAY, size=9)
        else:
            base_fill = None
            base_font = dict(size=9)

        ws.row_dimensions[xrow].height = 15 if level == 3 else 17

        for col_idx, (grp, sub) in enumerate(cols):
            xcol = col_idx + 1
            val  = row_data.iloc[col_idx]
            cell = ws.cell(row=xrow, column=xcol)
            cell.border = thin_border

            if base_fill:
                cell.fill = base_fill

            if grp == "Segmento":
                # Limpa nbsp; usa indent do openpyxl
                clean = str(val).replace("\u00a0", "") if pd.notna(val) else ""
                cell.value = clean
                indent = max(0, level) if not is_ipca else 0
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
                cell.font = Font(**base_font)

            elif grp in (_G_MEN, _G_ACUM):
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = '#,##0.0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(**base_font)

            elif grp in (_G_VAR, _G_CRES):
                is_part = (sub == "Part. % crescimento")
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = '0.0"%"' if is_part else '+0.0"%";-0.0"%";"0%"'
                    color_val = C_TEAL if float(val) >= 0 else C_RED
                    cell.font = Font(**{**base_font, "color": color_val})
                else:
                    cell.font = Font(**base_font)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            else:
                cell.value = val if pd.notna(val) else None
                cell.font = Font(**base_font)

    # ── Larguras de coluna ───────────────────────────────────────────────────
    for col_idx, (grp, sub) in enumerate(cols):
        letter = get_column_letter(col_idx + 1)
        if grp == "Segmento":
            ws.column_dimensions[letter].width = 34
        elif grp in (_G_MEN, _G_ACUM):
            ws.column_dimensions[letter].width = 11
        else:
            ws.column_dimensions[letter].width = 10

    ws.freeze_panes = "B3"

    wb.save(buf)
    return buf.getvalue()


def export_sinistralidade_excel(
    df_flat: "pd.DataFrame",
    levels: list,
    mlbls: list,
    ref_label: str,
    has_yoy: bool = True,
    sheet_name: str = "Sinistralidade",
) -> bytes:
    """
    Exporta tabela de sinistralidade para Excel com formatação CNseg.
    df_flat : DataFrame com colunas [Nome, *mlbls, M/M (p.p.), A/A (p.p.),
              Acum. Jan–{ref} (%), [Acum. A/A (p.p.)]]
    levels  : lista de int por linha (0=segmento, 1=grupo1, 2=grupo2)
    mlbls   : rótulos dos meses (5 itens)
    ref_label: rótulo do período de referência (ex. "Jan/2026")
    has_yoy : se existe coluna A/A acumulada
    """
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_DARK       = "403D4A"
    C_DARK2      = "5A5761"
    C_ORANGE_BG  = "FEF3E8"
    C_TEAL       = "20A787"
    C_RED        = "F44949"
    C_GRAY       = "A09FA4"
    WHITE        = "FFFFFF"

    # Grupos de colunas (MultiIndex)
    grp_tuples = [("Segmento", "")]
    for lbl in mlbls:
        grp_tuples.append(("Sinistralidade (%)", lbl))
    grp_tuples.append(("Variação (p.p.)", "M/M"))
    grp_tuples.append(("Variação (p.p.)", "A/A"))
    grp_tuples.append(("Acumulado", f"Jan–{ref_label}"))
    if has_yoy:
        grp_tuples.append(("Acumulado", "A/A (p.p.)"))

    buf = _io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = sheet_name

    thin_side   = Side(style="thin",   color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    # ── Cabeçalho linha 1: grupos mesclados ─────────────────────────────────
    fill1   = PatternFill(fill_type="solid", fgColor=C_DARK)
    font1   = Font(bold=True, color=WHITE, size=10)
    align_c = Alignment(horizontal="center", vertical="center")

    grp_spans: list = []
    prev_grp, start_i = None, 0
    for i, (g, _) in enumerate(grp_tuples):
        if g != prev_grp:
            if prev_grp is not None:
                grp_spans.append((prev_grp, start_i, i - 1))
            prev_grp, start_i = g, i
    grp_spans.append((prev_grp, start_i, len(grp_tuples) - 1))

    grp_border_side = Side(style="medium", color=WHITE)
    for grp, ci_start, ci_end in grp_spans:
        sc, ec = ci_start + 1, ci_end + 1
        cell = ws.cell(row=1, column=sc, value=grp if grp != "Segmento" else "")
        cell.font      = font1
        cell.fill      = fill1
        cell.alignment = align_c
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        for c in range(sc, ec + 1):
            ws.cell(row=1, column=c).fill = fill1
        ws.cell(row=1, column=sc).border = Border(left=grp_border_side)
        ws.cell(row=1, column=ec).border = Border(right=grp_border_side)

    # ── Cabeçalho linha 2: sub-colunas ──────────────────────────────────────
    fill2    = PatternFill(fill_type="solid", fgColor=C_DARK2)
    font2    = Font(bold=True, color=WHITE, size=9)
    align_c2 = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (grp, sub) in enumerate(grp_tuples):
        cell = ws.cell(row=2, column=i + 1, value=sub if sub else grp)
        cell.font      = font2
        cell.fill      = fill2
        cell.alignment = align_c2
        cell.border    = thin_border

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # ── Linhas de dados ──────────────────────────────────────────────────────
    _G_SR   = "Sinistralidade (%)"
    _G_VAR  = "Variação (p.p.)"
    _G_ACUM = "Acumulado"
    _acum_sr_sub = f"Jan–{ref_label}"

    for row_idx, row_data in df_flat.iterrows():
        xrow  = row_idx + 3
        level = levels[row_idx] if row_idx < len(levels) else 0

        if level == 0:
            base_fill = PatternFill(fill_type="solid", fgColor=C_ORANGE_BG)
            base_font = dict(bold=True, size=10)
        elif level == 1:
            base_fill = None
            base_font = dict(bold=True, size=9)
        elif level == 2:
            base_fill = None
            base_font = dict(color=C_GRAY, size=9)
        else:
            base_fill = None
            base_font = dict(size=9)

        ws.row_dimensions[xrow].height = 15 if level == 2 else 17

        for col_idx, (grp, sub) in enumerate(grp_tuples):
            xcol = col_idx + 1
            val  = row_data.iloc[col_idx]
            cell = ws.cell(row=xrow, column=xcol)
            cell.border = thin_border
            if base_fill:
                cell.fill = base_fill

            if grp == "Segmento":
                clean = str(val).replace("\u00a0", "") if pd.notna(val) else ""
                cell.value     = clean
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           indent=max(0, level))
                cell.font = Font(**base_font)

            elif grp == _G_SR or (grp == _G_ACUM and sub == _acum_sr_sub):
                # Sinistralidade % — formato sem sinal
                if pd.notna(val):
                    cell.value         = float(val)
                    cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(**base_font)

            elif grp in (_G_VAR, _G_ACUM):
                # p.p. — cor pelo sinal (vermelho=alta, teal=queda)
                if pd.notna(val):
                    cell.value         = float(val)
                    cell.number_format = '+0.0;-0.0;"0,0"'
                    color_val = C_RED if float(val) > 0 else C_TEAL
                    cell.font = Font(**{**base_font, "color": color_val})
                else:
                    cell.font = Font(**base_font)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            else:
                cell.value = val if pd.notna(val) else None
                cell.font  = Font(**base_font)

    # ── Larguras de coluna ───────────────────────────────────────────────────
    for col_idx, (grp, _) in enumerate(grp_tuples):
        letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[letter].width = 34 if grp == "Segmento" else 10

    ws.freeze_panes = "B3"
    wb.save(buf)
    return buf.getvalue()


def export_divergencias_excel(df: pd.DataFrame) -> bytes:
    """
    Exporta tabela de divergências para Excel com estilo CNseg.
    Header escuro, auto-filtro, "Diferença Relativa (%)" em vermelho,
    "Tipo" colorido por categoria, linha de total no rodapé.
    """
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_DARK      = "403D4A"
    C_ORANGE_BG = "FEF3E8"
    C_RED       = "F44949"
    C_YELLOW    = "F5A623"
    WHITE       = "FFFFFF"

    TIPO_CORES = {
        "Valores diferentes":          C_RED,
        "Período ausente na API":       C_YELLOW,
        "Período ausente no Cognos":    C_YELLOW,
    }

    thin_side   = Side(style="thin", color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    COL_WIDTHS = {
        "Fonte": 10, "Indicador": 42, "Ano": 7, "Mês": 6,
        "Valor API": 14, "Valor Cognos": 14,
        "Diferença Absoluta": 16, "Diferença Relativa (%)": 16, "Tipo": 26,
    }
    NUM_COLS = {"Valor API", "Valor Cognos", "Diferença Absoluta", "Diferença Relativa (%)"}
    INT_COLS = {"Ano", "Mês"}

    buf = _io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Divergências"

    cols = df.columns.tolist()
    n_cols = len(cols)

    # ── Header ────────────────────────────────────────────────────────────────
    fill_h  = PatternFill(fill_type="solid", fgColor=C_DARK)
    font_h  = Font(bold=True, color=WHITE, size=10)
    align_h = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = font_h, fill_h, align_h, thin_border
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)

    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Dados ─────────────────────────────────────────────────────────────────
    fill_alt = PatternFill(fill_type="solid", fgColor=C_ORANGE_BG)

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[ri].height = 15
        tipo_val = str(row.get("Tipo", "")) if "Tipo" in row.index else ""
        use_alt  = tipo_val in TIPO_CORES

        for ci, col in enumerate(cols, start=1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin_border

            if use_alt:
                cell.fill = fill_alt

            if col in INT_COLS:
                cell.value = int(val) if pd.notna(val) else None
                cell.font  = Font(size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col in NUM_COLS:
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = "0.0000"
                    if col == "Diferença Relativa (%)":
                        cell.font = Font(size=9, bold=True, color=C_RED)
                    else:
                        cell.font = Font(size=9)
                else:
                    cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            elif col == "Tipo":
                cell.value = str(val) if pd.notna(val) else ""
                cor = TIPO_CORES.get(str(val), C_DARK)
                cell.font  = Font(size=9, bold=True, color=cor)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.font  = Font(size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=(col == "Indicador"))

    # ── Linha de total ────────────────────────────────────────────────────────
    total_row = len(df) + 2
    ws.row_dimensions[total_row].height = 17
    fill_total = PatternFill(fill_type="solid", fgColor="EDE9F0")
    font_total = Font(bold=True, size=9, color=C_DARK)

    n_div   = len(df)
    por_tipo = df["Tipo"].value_counts().to_dict() if "Tipo" in df.columns else {}
    resumo  = "  ·  ".join(f"{v} {k}" for k, v in por_tipo.items()) if por_tipo else ""
    total_txt = f"Total: {n_div} divergência(s)   {resumo}"

    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=ci)
        cell.fill   = fill_total
        cell.border = thin_border
        cell.font   = font_total
    ws.cell(row=total_row, column=1).value = total_txt
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=n_cols)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()


def export_dados_api_excel(
    df: pd.DataFrame,
    fonte_label: str = None,
    tab_color: str = None,
    periodo_label: str = None,
) -> bytes:
    """
    Exporta dados brutos da API para Excel com estilo CNseg.
    Linha 1: metadados (fonte, período, contagens, timestamp).
    Linha 2: cabeçalho estilizado com auto-filtro.
    Dados a partir da linha 3, sem estilo por célula (performance).
    Tab colorida conforme a fonte.
    """
    import io as _io
    import datetime
    from zoneinfo import ZoneInfo
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_DARK  = "403D4A"
    C_DARK2 = "5A5761"
    C_GRAY  = "A09FA4"
    WHITE   = "FFFFFF"

    thin_side   = Side(style="thin", color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    cols   = df.columns.tolist()
    n_cols = len(cols)
    n_ind  = n_cols - 2  # exclui Ano e Mês
    n_per  = len(df)
    agora  = datetime.datetime.now(tz=ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M")

    partes = []
    if fonte_label:
        partes.append(fonte_label)
    if periodo_label:
        partes.append(periodo_label)
    partes += [f"{n_per} períodos", f"{n_ind} indicadores", f"Gerado em {agora}"]
    metadata_txt = "   ·   ".join(partes)

    buf = _io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = fonte_label or "Dados API"

    if tab_color:
        ws.sheet_properties.tabColor = tab_color.lstrip("#")

    # ── Linha de metadados ────────────────────────────────────────────────────
    meta_cell = ws.cell(row=1, column=1, value=metadata_txt)
    meta_cell.font      = Font(italic=True, size=8, color=C_GRAY)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 16

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    font_h  = Font(bold=True, color=WHITE, size=9)
    align_h = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, start=1):
        is_key = col in ("Ano", "Mês")
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = font_h
        cell.fill      = PatternFill(fill_type="solid", fgColor=C_DARK if is_key else C_DARK2)
        cell.alignment = align_h
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = 7 if is_key else 16

    ws.row_dimensions[2].height = 32
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"

    # ── Dados (append direto, sem estilo por célula) ──────────────────────────
    for row in df.itertuples(index=False):
        ws.append(list(row))

    ws.freeze_panes = "C3"
    wb.save(buf)
    return buf.getvalue()


def make_demonst_contabil_table(ds2_q: "pd.DataFrame", q_ref: int):
    """
    Builds quarterly demonst_contabil table with MultiIndex columns.

    ds2_q: DataFrame with columns: qcode (YYYYQ int), trim_label (str), modalidade, + financial cols
           Values in raw R$. Function divides by 1e6 for display.
    q_ref: reference qcode (e.g., 20253 for 3T2025)

    Returns (df, levels) where:
      df: MultiIndex DataFrame (columns = MultiIndex)
      levels: list[int] per row  (0 = modalidade header, 1 = rubrica)
    """
    import pandas as pd

    _RUBRICA_LABELS = {
        "arrecadacao":   "Arrecadação",
        "sinistro":      "Sinistros/Benef./Indeniz.",
        "desp_admin":    "Despesas Administrativas",
        "desp_tributo":  "Despesas com Tributos",
        "resultado_fin": "Resultado Financeiro",
        "lucro_liquido": "Lucro Líquido",
        "pl":            "Patrimônio Líquido",
        "ativo":         "Ativo",
        "provisao":      "Provisão",
    }
    _SIN_OVERRIDE = {
        "Capitalização":      "Resgates e Sorteios",
        "Saúde Suplementar":  "Eventos Indenizáveis",
    }

    all_qcodes = sorted(ds2_q["qcode"].unique())
    q_idx = next((i for i, q in enumerate(all_qcodes) if q == q_ref), len(all_qcodes) - 1)
    q_last6 = all_qcodes[max(0, q_idx - 5): q_idx + 1]

    def _ql(qc):
        return f"{qc % 10}T{qc // 10}"

    q1 = q_last6[-1] if len(q_last6) >= 1 else None
    q2 = q_last6[-2] if len(q_last6) >= 2 else None
    q3 = q_last6[-3] if len(q_last6) >= 3 else None

    def _yoy(qc):
        return (qc // 10 - 1) * 10 + (qc % 10) if qc else None

    q_yoy1 = _yoy(q1)
    q_yoy2 = _yoy(q2)
    q_yoy1 = q_yoy1 if q_yoy1 in all_qcodes else None
    q_yoy2 = q_yoy2 if q_yoy2 in all_qcodes else None

    q_year = q1 // 10 if q1 else None
    q_num  = q1 % 10  if q1 else None
    q_prev_year = (q_year - 1) if q_year else None

    q_ytd_curr = [q for q in all_qcodes if q // 10 == q_year      and q % 10 <= q_num] if q_year else []
    q_ytd_prev = [q for q in all_qcodes if q // 10 == q_prev_year  and q % 10 <= q_num] if q_prev_year else []

    q_labels = [_ql(q) for q in q_last6]

    def _vl(a, b):
        return f"{_ql(a)}/{_ql(b)}" if a and b else "—"

    top = (
        ["Rubrica"] +
        ["Trimestral (R$ mi)"] * len(q_labels) +
        ["Variação Nominal %"] * 4 +
        [f"Acum. até {q_num}º tri (R$ mi)"] * 2 +
        ["Var. Acum. %"]
    )
    sub = (
        [""] +
        q_labels +
        [_vl(q1, q2), _vl(q2, q3), _vl(q1, q_yoy1), _vl(q2, q_yoy2)] +
        [str(q_prev_year), str(q_year)] +
        [f"{q_year}/{q_prev_year}"]
    )
    mi = pd.MultiIndex.from_arrays([top, sub])

    financial_cols = [c for c in _RUBRICA_LABELS if c in ds2_q.columns]
    mods = sorted(ds2_q["modalidade"].dropna().unique())

    rows_data = []
    levels = []

    for mod in mods:
        mod_df = ds2_q[ds2_q["modalidade"] == mod]

        def gq(qcode, col):
            if qcode is None or col not in mod_df.columns:
                return None
            r = mod_df[mod_df["qcode"] == qcode][col].dropna()
            return float(r.sum()) if not r.empty else None

        _STOCK_COLS_SET = {"pl", "ativo", "provisao"}

        def gytd(qcodes, col):
            if not qcodes or col not in mod_df.columns:
                return None
            r = mod_df[mod_df["qcode"].isin(qcodes)].sort_values("qcode")[col].dropna()
            if r.empty:
                return None
            # Estoques (PL, Ativo, Provisão): saldo do último trimestre do período
            return float(r.iloc[-1]) if col in _STOCK_COLS_SET else float(r.sum())

        def pct(a, b):
            if a is None or b is None or b == 0:
                return None
            return (a - b) / abs(b) * 100

        # Modalidade header row (all None except label)
        rows_data.append([mod] + [None] * (len(q_labels) + 4 + 2 + 1))
        levels.append(0)

        for col_key in financial_cols:
            sin_lbl = _SIN_OVERRIDE.get(mod, _RUBRICA_LABELS["sinistro"]) if col_key == "sinistro" else _RUBRICA_LABELS[col_key]
            lbl = "\u00a0\u00a0" + sin_lbl

            q_vals = [gq(q, col_key) for q in q_last6]
            q_vals_mi = [(v / 1e6 if v is not None else None) for v in q_vals]

            v1 = gq(q1, col_key)
            v2 = gq(q2, col_key)
            v3 = gq(q3, col_key)
            vy1 = gq(q_yoy1, col_key)
            vy2 = gq(q_yoy2, col_key)

            ytd_c = gytd(q_ytd_curr, col_key)
            ytd_p = gytd(q_ytd_prev, col_key)

            row = (
                [lbl] +
                q_vals_mi +
                [pct(v1, v2), pct(v2, v3), pct(v1, vy1), pct(v2, vy2)] +
                [(ytd_p / 1e6 if ytd_p is not None else None),
                 (ytd_c / 1e6 if ytd_c is not None else None)] +
                [pct(ytd_c, ytd_p)]
            )
            rows_data.append(row)
            levels.append(1)

    df = pd.DataFrame(rows_data, columns=mi)
    return df, levels


def style_demonst_contabil_table(df: "pd.DataFrame", levels: list) -> "pd.io.formats.style.Styler":
    """Styles the demonst_contabil quarterly table returned by make_demonst_contabil_table."""
    from ui.colors import CNSEG_ORANGE

    def _row_style(row):
        idx = row.name
        lvl = levels[idx] if idx < len(levels) else 1
        if lvl == 0:
            return [f"font-weight:bold; background-color:rgba(247,135,31,0.18)"] * len(row)
        return [""] * len(row)

    styled = df.style.apply(_row_style, axis=1)

    for col in df.columns:
        top, sub = col
        if top == "Rubrica":
            continue
        if "%" in top or "%" in sub or top.startswith("Var"):
            styled = styled.format(lambda v: _br(v, "+.1f") + "%", subset=[(top, sub)], na_rep="—")
        else:
            styled = styled.format(lambda v: _br(v, ",.1f"), subset=[(top, sub)], na_rep="—")

    return styled


def export_dc_excel(tbl: "pd.DataFrame", levels: list, sheet_name: str = "Demonst. Contábil") -> bytes:
    """Exporta make_demonst_contabil_table para Excel formatado."""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_DARK      = "403D4A"
    C_DARK2     = "5A5761"
    C_ORANGE_BG = "FDE0B8"
    C_TEAL      = "20A787"
    C_RED       = "F44949"
    WHITE       = "FFFFFF"

    buf = _io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = sheet_name

    cols = tbl.columns.tolist()
    thin = Side(style="thin", color="DDDDDD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row 1: grupos mesclados ──────────────────────────────────────
    fill1 = PatternFill(fill_type="solid", fgColor=C_DARK)
    font1 = Font(bold=True, color=WHITE, size=10)
    ac    = Alignment(horizontal="center", vertical="center")

    grp_spans: list = []
    prev_g, start_i = None, 0
    for i, (g, _) in enumerate(cols):
        if g != prev_g:
            if prev_g is not None:
                grp_spans.append((prev_g, start_i, i - 1))
            prev_g, start_i = g, i
    grp_spans.append((prev_g, start_i, len(cols) - 1))

    grp_border_side = Side(style="medium", color=WHITE)
    for grp, ci0, ci1 in grp_spans:
        sc, ec = ci0 + 1, ci1 + 1
        cell = ws.cell(row=1, column=sc, value="" if grp == "Rubrica" else grp)
        cell.font, cell.fill, cell.alignment = font1, fill1, ac
        if sc != ec:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        for c in range(sc, ec + 1):
            ws.cell(row=1, column=c).fill = fill1
        ws.cell(row=1, column=sc).border = Border(left=grp_border_side)
        ws.cell(row=1, column=ec).border = Border(right=grp_border_side)

    # ── Header row 2: sub-colunas ────────────────────────────────────────────
    fill2 = PatternFill(fill_type="solid", fgColor=C_DARK2)
    font2 = Font(bold=True, color=WHITE, size=9)
    ac2   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (grp, sub) in enumerate(cols):
        cell = ws.cell(row=2, column=i + 1, value=sub if sub else grp)
        cell.font, cell.fill, cell.alignment, cell.border = font2, fill2, ac2, brd

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36

    # ── Dados ────────────────────────────────────────────────────────────────
    for row_idx, row_data in tbl.iterrows():
        xrow   = row_idx + 3
        level  = levels[row_idx] if row_idx < len(levels) else 1
        is_mod = (level == 0)
        bg = PatternFill(fill_type="solid", fgColor=C_ORANGE_BG if is_mod else "FFFFFF")
        base_font_kw = dict(bold=is_mod, color=C_DARK, size=10 if is_mod else 9)

        for ci, ((grp, sub), val) in enumerate(zip(cols, row_data)):
            cell = ws.cell(row=xrow, column=ci + 1)
            cell.border, cell.fill = brd, bg
            if ci == 0:
                cell.value     = str(val) if val is not None else ""
                cell.font      = Font(**base_font_kw)
                cell.alignment = Alignment(horizontal="left")
            elif val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = None
            else:
                is_pct = "%" in grp or "Var." in grp
                cell.value = float(val)
                if is_pct:
                    cell.number_format = '0.0"%"'
                    c_val = C_TEAL if float(val) >= 0 else C_RED
                    cell.font = Font(bold=is_mod, color=c_val, size=base_font_kw["size"])
                else:
                    cell.number_format = "#,##0"
                    cell.font = Font(**base_font_kw)
                cell.alignment = Alignment(horizontal="right")

        ws.row_dimensions[xrow].height = 18 if is_mod else 16

    # ── Larguras e freeze ────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 34
    for ci in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "B3"

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
